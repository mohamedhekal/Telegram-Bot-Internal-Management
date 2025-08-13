#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
تحويل البيانات الموجودة في قاعدة البيانات إلى الشكل الجديد المطلوب
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
from database_manager import DatabaseManager

# قاموس مفاتيح المحافظات
GOVERNORATE_CODES = {
    'بغداد': 'BGD',
    'الناصرية ذي قار': 'NAS',
    'ديالى': 'DYL',
    'الكوت واسط': 'KOT',
    'كربلاء': 'KRB',
    'دهوك': 'DOH',
    'بابل الحلة': 'BBL',
    'النجف': 'NJF',
    'البصرة': 'BAS',
    'اربيل': 'ARB',
    'كركوك': 'KRK',
    'السليمانيه': 'SMH',
    'صلاح الدين': 'SAH',
    'الانبار رمادي': 'ANB',
    'السماوة المثنى': 'SAM',
    'موصل': 'MOS',
    'الديوانية': 'DWN',
    'العمارة ميسان': 'AMA'
}

def get_governorate_code(governorate_name):
    """الحصول على شفرة المحافظة من الاسم"""
    # البحث المباشر
    if governorate_name in GOVERNORATE_CODES:
        return GOVERNORATE_CODES[governorate_name]
    
    # البحث الجزئي
    for name, code in GOVERNORATE_CODES.items():
        if governorate_name in name or name in governorate_name:
            return code
    
    # إذا لم يتم العثور على تطابق، إرجاع القيمة الأصلية
    return governorate_name

def convert_existing_data_to_new_format(days=30):
    """تحويل البيانات الموجودة إلى الشكل الجديد المطلوب"""
    
    print(f"🔄 تحويل البيانات للآخر {days} يوم...")
    
    # الحصول على البيانات من قاعدة البيانات
    db_manager = DatabaseManager()
    df = db_manager.get_all_invoices_for_shipping(days)
    
    if df is None or df.empty:
        print("❌ لا توجد بيانات للتحويل")
        return None
    
    print(f"📊 تم العثور على {len(df)} طلب")
    
    # تحويل البيانات إلى الشكل الجديد
    converted_data = []
    
    for index, row in df.iterrows():
        # تحويل البيانات
        converted_row = {
            'ملاحظات': row.get('notes', ''),
            'عدد القطع\nأجباري': row.get('quantity', 0),
            'يحتوي على ارجاع بضاعة؟': 'لا',  # افتراضي
            'هاتف المستلم\nأجباري 11 خانة': row.get('client_phone', ''),
            'تفاصيل العنوان\nأجباري': row.get('nearest_point', ''),
            'شفرة المحافظة\nأجباري': get_governorate_code(row.get('governorate', '')),
            'أسم المستلم': row.get('client_name', ''),
            'المبلغ عراقي\nكامل بالالاف .\nفي حال عدم توفره سيعتبر 0': row.get('total_sales', 0)
        }
        converted_data.append(converted_row)
    
    # إنشاء DataFrame جديد
    new_df = pd.DataFrame(converted_data)
    
    # إنشاء ملف الإكسل
    filename = f"طلبات_محولة_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # كتابة البيانات الرئيسية
        new_df.to_excel(writer, sheet_name='طلبات الشحن', index=False)
        
        # الحصول على ورقة العمل
        workbook = writer.book
        worksheet = writer.sheets['طلبات الشحن']
        
        # تنسيق العناوين
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # تنسيق حدود العناوين
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # تطبيق التنسيق على العناوين
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # تعديل عرض الأعمدة
        column_widths = {
            'A': 20,  # ملاحظات
            'B': 15,  # عدد القطع
            'C': 25,  # ارجاع بضاعة
            'D': 25,  # هاتف المستلم
            'E': 30,  # تفاصيل العنوان
            'F': 20,  # شفرة المحافظة
            'G': 20,  # أسم المستلم
            'H': 35   # المبلغ عراقي
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # تنسيق البيانات
        data_alignment = Alignment(horizontal="center", vertical="center")
        data_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # تطبيق التنسيق على البيانات
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = data_border
        
        # إنشاء ورقة مفاتيح المحافظات
        governorate_sheet = workbook.create_sheet("مفاتيح المحافظات")
        
        # إضافة مفاتيح المحافظات
        governorate_data = []
        for governorate, code in GOVERNORATE_CODES.items():
            governorate_data.append([governorate, code])
        
        governorate_df = pd.DataFrame(governorate_data, columns=['المحافظة', 'الشفرة'])
        governorate_df.to_excel(writer, sheet_name='مفاتيح المحافظات', index=False)
        
        # تنسيق ورقة مفاتيح المحافظات
        gov_worksheet = writer.sheets['مفاتيح المحافظات']
        
        # تنسيق العناوين
        for cell in gov_worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # تعديل عرض الأعمدة
        gov_worksheet.column_dimensions['A'].width = 25
        gov_worksheet.column_dimensions['B'].width = 15
        
        # تنسيق البيانات
        for row in gov_worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = data_border
        
        # إنشاء ورقة ملخص التحويل
        summary_sheet = workbook.create_sheet("ملخص التحويل")
        
        summary_data = [
            ['ملخص تحويل البيانات'],
            [''],
            ['تاريخ التحويل:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['عدد الطلبات المحولة:', len(converted_data)],
            ['فترة البيانات:', f'آخر {days} يوم'],
            [''],
            ['تفاصيل التحويل:'],
            ['- تم تحويل البيانات من قاعدة البيانات إلى الشكل الجديد'],
            ['- تم إضافة شفرات المحافظات تلقائياً'],
            ['- تم تعيين "لا" كقيمة افتراضية لعمود "يحتوي على ارجاع بضاعة؟"'],
            ['- تم الحفاظ على جميع البيانات الأصلية'],
            [''],
            ['ملاحظات:'],
            ['- راجع البيانات المحولة وتأكد من صحتها'],
            ['- يمكن تعديل البيانات يدوياً إذا لزم الأمر'],
            ['- استخدم ورقة "مفاتيح المحافظات" للتحقق من الشفرات']
        ]
        
        for i, summary in enumerate(summary_data, 1):
            summary_sheet[f'A{i}'] = summary[0]
            if i == 1:  # العنوان الرئيسي
                summary_sheet[f'A{i}'].font = Font(bold=True, size=14)
            elif i in [7, 13]:  # العناوين الفرعية
                summary_sheet[f'A{i}'].font = Font(bold=True, size=12)
        
        # تعديل عرض عمود الملخص
        summary_sheet.column_dimensions['A'].width = 60
    
    print(f"✅ تم إنشاء ملف البيانات المحولة بنجاح: {filename}")
    print(f"📊 تم تحويل {len(converted_data)} طلب")
    
    return filename

def show_conversion_stats():
    """عرض إحصائيات التحويل"""
    
    db_manager = DatabaseManager()
    
    # إحصائيات عامة
    print("📊 إحصائيات قاعدة البيانات:")
    print("=" * 40)
    
    # إجمالي الفواتير
    conn = db_manager.get_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) FROM invoices")
    total_invoices = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM invoices WHERE created_at >= datetime('now', '-1 day')")
    today_invoices = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM invoices WHERE created_at >= datetime('now', '-7 days')")
    week_invoices = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM invoices WHERE created_at >= datetime('now', '-30 days')")
    month_invoices = cursor.fetchone()[0]
    
    conn.close()
    
    print(f"إجمالي الفواتير: {total_invoices}")
    print(f"فواتير اليوم: {today_invoices}")
    print(f"فواتير الأسبوع: {week_invoices}")
    print(f"فواتير الشهر: {month_invoices}")
    
    # إحصائيات المحافظات
    print("\n🏛️ إحصائيات المحافظات:")
    print("-" * 30)
    
    df = db_manager.get_all_invoices_for_shipping(365)  # سنة كاملة
    if df is not None and not df.empty:
        governorate_stats = df['governorate'].value_counts()
        for governorate, count in governorate_stats.head(10).items():
            code = get_governorate_code(governorate)
            print(f"{governorate} ({code}): {count} طلب")

if __name__ == "__main__":
    print("🔄 تحويل البيانات الموجودة إلى الشكل الجديد...")
    print("=" * 60)
    
    # عرض الإحصائيات
    show_conversion_stats()
    
    print("\n" + "=" * 60)
    
    # تحويل البيانات للآخر 30 يوم
    converted_file = convert_existing_data_to_new_format(days=30)
    
    if converted_file:
        print(f"\n📁 الملف المنشأ: {converted_file}")
        print("🎯 يمكنك الآن استخدام هذا الملف للشحن!")
    else:
        print("\n❌ لم يتم إنشاء أي ملف")
