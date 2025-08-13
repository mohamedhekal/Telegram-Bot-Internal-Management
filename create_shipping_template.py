#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
إنشاء ملف إكسل قالب الشحن بالشكل المطلوب
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

# قاموس مفاتيح المحافظات
GOVERNORATE_CODES = {
    'بغداد': 'BGD',
    'الناصرية ذي قار': 'NAS',
    'ديالى': 'DYL',
    'الكوت واسط': 'KOT',
    'كربلاء': 'KRB',
    'دهوك': 'DOH',
    'بابل الحلة': 'BBL',
    'النجف': 'NJF',
    'البصرة': 'BAS',
    'اربيل': 'ARB',
    'كركوك': 'KRK',
    'السليمانيه': 'SMH',
    'صلاح الدين': 'SAH',
    'الانبار رمادي': 'ANB',
    'السماوة المثنى': 'SAM',
    'موصل': 'MOS',
    'الديوانية': 'DWN',
    'العمارة ميسان': 'AMA'
}

def create_shipping_template():
    """إنشاء ملف إكسل قالب الشحن بالشكل المطلوب"""
    
    # إنشاء DataFrame فارغ مع الأعمدة المطلوبة
    columns = [
        'ملاحظات',
        'عدد القطع\nأجباري',
        'يحتوي على ارجاع بضاعة؟',
        'هاتف المستلم\nأجباري 11 خانة',
        'تفاصيل العنوان\nأجباري',
        'شفرة المحافظة\nأجباري',
        'أسم المستلم',
        'المبلغ عراقي\nكامل بالالاف .\nفي حال عدم توفره سيعتبر 0'
    ]
    
    # إنشاء DataFrame فارغ
    df = pd.DataFrame(columns=columns)
    
    # إضافة صف تجريبي للعرض
    sample_data = {
        'ملاحظات': 'ملاحظات تجريبية',
        'عدد القطع\nأجباري': 5,
        'يحتوي على ارجاع بضاعة؟': 'لا',
        'هاتف المستلم\nأجباري 11 خانة': '07801234567',
        'تفاصيل العنوان\nأجباري': 'شارع الرشيد - بغداد',
        'شفرة المحافظة\nأجباري': 'BGD',
        'أسم المستلم': 'أحمد محمد',
        'المبلغ عراقي\nكامل بالالاف .\nفي حال عدم توفره سيعتبر 0': 50000
    }
    
    df = pd.concat([df, pd.DataFrame([sample_data])], ignore_index=True)
    
    # إنشاء ملف الإكسل
    filename = f"قالب_الشحن_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # كتابة البيانات الرئيسية
        df.to_excel(writer, sheet_name='طلبات الشحن', index=False)
        
        # الحصول على ورقة العمل
        workbook = writer.book
        worksheet = writer.sheets['طلبات الشحن']
        
        # تنسيق العناوين
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # تنسيق حدود العناوين
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # تطبيق التنسيق على العناوين
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # تعديل عرض الأعمدة
        column_widths = {
            'A': 20,  # ملاحظات
            'B': 15,  # عدد القطع
            'C': 25,  # ارجاع بضاعة
            'D': 25,  # هاتف المستلم
            'E': 30,  # تفاصيل العنوان
            'F': 20,  # شفرة المحافظة
            'G': 20,  # أسم المستلم
            'H': 35   # المبلغ عراقي
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # تنسيق البيانات
        data_alignment = Alignment(horizontal="center", vertical="center")
        data_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # تطبيق التنسيق على البيانات
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = data_border
        
        # إنشاء ورقة مفاتيح المحافظات
        governorate_sheet = workbook.create_sheet("مفاتيح المحافظات")
        
        # إضافة مفاتيح المحافظات
        governorate_data = []
        for governorate, code in GOVERNORATE_CODES.items():
            governorate_data.append([governorate, code])
        
        governorate_df = pd.DataFrame(governorate_data, columns=['المحافظة', 'الشفرة'])
        governorate_df.to_excel(writer, sheet_name='مفاتيح المحافظات', index=False)
        
        # تنسيق ورقة مفاتيح المحافظات
        gov_worksheet = writer.sheets['مفاتيح المحافظات']
        
        # تنسيق العناوين
        for cell in gov_worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # تعديل عرض الأعمدة
        gov_worksheet.column_dimensions['A'].width = 25
        gov_worksheet.column_dimensions['B'].width = 15
        
        # تنسيق البيانات
        for row in gov_worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = data_border
        
        # إضافة تعليمات الاستخدام
        instructions_sheet = workbook.create_sheet("تعليمات الاستخدام")
        
        instructions = [
            ['تعليمات استخدام قالب الشحن'],
            [''],
            ['الأعمدة المطلوبة:'],
            ['1. ملاحظات - يمكن تركها فارغة'],
            ['2. عدد القطع - أجباري، يجب إدخال رقم صحيح'],
            ['3. يحتوي على ارجاع بضاعة؟ - أجباري، "نعم" أو "لا"'],
            ['4. هاتف المستلم - أجباري، 11 خانة رقم'],
            ['5. تفاصيل العنوان - أجباري، عنوان مفصل'],
            ['6. شفرة المحافظة - أجباري، راجع ورقة مفاتيح المحافظات'],
            ['7. أسم المستلم - أجباري، اسم المستلم الكامل'],
            ['8. المبلغ عراقي - أجباري، المبلغ بالدينار العراقي (بالآلاف)'],
            [''],
            ['ملاحظات مهمة:'],
            ['- في حال عدم توفر المبلغ، سيتم اعتباره 0'],
            ['- يجب التأكد من صحة رقم الهاتف (11 خانة)'],
            ['- شفرة المحافظة يجب أن تكون من القائمة المحددة'],
            ['- يمكن إضافة ملاحظات إضافية في عمود الملاحظات']
        ]
        
        for i, instruction in enumerate(instructions, 1):
            instructions_sheet[f'A{i}'] = instruction[0]
            if i == 1:  # العنوان الرئيسي
                instructions_sheet[f'A{i}'].font = Font(bold=True, size=14)
            elif i in [3, 12]:  # العناوين الفرعية
                instructions_sheet[f'A{i}'].font = Font(bold=True, size=12)
        
        # تعديل عرض عمود التعليمات
        instructions_sheet.column_dimensions['A'].width = 60
    
    print(f"✅ تم إنشاء ملف قالب الشحن بنجاح: {filename}")
    print(f"📋 يحتوي الملف على:")
    print(f"   - ورقة طلبات الشحن مع الأعمدة المطلوبة")
    print(f"   - ورقة مفاتيح المحافظات ({len(GOVERNORATE_CODES)} محافظة)")
    print(f"   - ورقة تعليمات الاستخدام")
    
    return filename

def create_empty_shipping_template():
    """إنشاء ملف إكسل فارغ للشحن بدون بيانات تجريبية"""
    
    # إنشاء DataFrame فارغ مع الأعمدة المطلوبة
    columns = [
        'ملاحظات',
        'عدد القطع\nأجباري',
        'يحتوي على ارجاع بضاعة؟',
        'هاتف المستلم\nأجباري 11 خانة',
        'تفاصيل العنوان\nأجباري',
        'شفرة المحافظة\nأجباري',
        'أسم المستلم',
        'المبلغ عراقي\nكامل بالالاف .\nفي حال عدم توفره سيعتبر 0'
    ]
    
    df = pd.DataFrame(columns=columns)
    
    # إنشاء ملف الإكسل
    filename = f"قالب_الشحن_فارغ_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # كتابة البيانات الرئيسية
        df.to_excel(writer, sheet_name='طلبات الشحن', index=False)
        
        # الحصول على ورقة العمل
        workbook = writer.book
        worksheet = writer.sheets['طلبات الشحن']
        
        # تنسيق العناوين
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # تنسيق حدود العناوين
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # تطبيق التنسيق على العناوين
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # تعديل عرض الأعمدة
        column_widths = {
            'A': 20,  # ملاحظات
            'B': 15,  # عدد القطع
            'C': 25,  # ارجاع بضاعة
            'D': 25,  # هاتف المستلم
            'E': 30,  # تفاصيل العنوان
            'F': 20,  # شفرة المحافظة
            'G': 20,  # أسم المستلم
            'H': 35   # المبلغ عراقي
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # إنشاء ورقة مفاتيح المحافظات
        governorate_sheet = workbook.create_sheet("مفاتيح المحافظات")
        
        # إضافة مفاتيح المحافظات
        governorate_data = []
        for governorate, code in GOVERNORATE_CODES.items():
            governorate_data.append([governorate, code])
        
        governorate_df = pd.DataFrame(governorate_data, columns=['المحافظة', 'الشفرة'])
        governorate_df.to_excel(writer, sheet_name='مفاتيح المحافظات', index=False)
        
        # تنسيق ورقة مفاتيح المحافظات
        gov_worksheet = writer.sheets['مفاتيح المحافظات']
        
        # تنسيق العناوين
        for cell in gov_worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # تعديل عرض الأعمدة
        gov_worksheet.column_dimensions['A'].width = 25
        gov_worksheet.column_dimensions['B'].width = 15
        
        # تنسيق البيانات
        data_alignment = Alignment(horizontal="center", vertical="center")
        data_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in gov_worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = data_border
    
    print(f"✅ تم إنشاء ملف قالب الشحن الفارغ بنجاح: {filename}")
    return filename

if __name__ == "__main__":
    print("🚀 إنشاء قوالب الشحن...")
    print("=" * 50)
    
    # إنشاء القالب مع بيانات تجريبية
    template_file = create_shipping_template()
    
    # إنشاء القالب الفارغ
    empty_template_file = create_empty_shipping_template()
    
    print("\n📁 الملفات المنشأة:")
    print(f"1. {template_file} - قالب مع بيانات تجريبية")
    print(f"2. {empty_template_file} - قالب فارغ للاستخدام")
    
    print("\n🎯 يمكنك الآن استخدام هذه القوالب لتحميل الطلبات!")
