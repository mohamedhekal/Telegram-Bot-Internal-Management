import sqlite3
import pandas as pd
from datetime import datetime, timedelta
import os
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import config

class DatabaseManager:
    def __init__(self, db_path="invoice_bot.db"):
        self.db_path = db_path
        self.init_database()
    
    def get_connection(self):
        """الحصول على اتصال قاعدة البيانات مع إعدادات محسنة"""
        conn = sqlite3.connect(self.db_path, timeout=30.0)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA cache_size=10000")
        conn.execute("PRAGMA temp_store=MEMORY")
        return conn
    
    def init_database(self):
        """تهيئة قاعدة البيانات"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # جدول الفواتير
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                receipt_number TEXT UNIQUE,
                employee_name TEXT,
                client_name TEXT,
                client_phone TEXT,
                governorate TEXT,
                nearest_point TEXT,
                quantity INTEGER,
                price REAL,
                total_sales REAL,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # جدول المستخدمين
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER UNIQUE,
                username TEXT,
                full_name TEXT,
                role TEXT DEFAULT 'employee',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # جدول كلمات مرور الموظفين
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS employee_passwords (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_name TEXT UNIQUE,
                password TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # جدول الإحصائيات اليومية
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS daily_stats (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_name TEXT,
                date DATE,
                total_orders INTEGER DEFAULT 0,
                total_quantity INTEGER DEFAULT 0,
                total_sales REAL DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(employee_name, date)
            )
        ''')
        
        # جدول المرتجعات
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS returns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER,
                receipt_number TEXT,
                employee_name TEXT,
                return_type TEXT CHECK(return_type IN ('full', 'partial')),
                returned_quantity INTEGER,
                returned_amount REAL,
                remaining_amount REAL,
                return_reason TEXT,
                processed_by TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (invoice_id) REFERENCES invoices (id)
            )
        ''')
        
        # جدول تتبع الطلبات المصدرة
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS exported_invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER,
                receipt_number TEXT,
                export_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                export_type TEXT,
                exported_by INTEGER,
                FOREIGN KEY (invoice_id) REFERENCES invoices (id)
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def add_invoice(self, invoice_data):
        """إضافة فاتورة جديدة"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO invoices (
                    receipt_number, employee_name, client_name, client_phone,
                    governorate, nearest_point, quantity, price, total_sales, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                invoice_data['receipt_number'],
                invoice_data['employee_name'],
                invoice_data['client_name'],
                invoice_data['client_phone'],
                invoice_data['governorate'],
                invoice_data['nearest_point'],
                invoice_data['quantity'],
                invoice_data['price'],
                invoice_data['total_sales'],
                invoice_data['notes']
            ))
            
            # تحديث الإحصائيات اليومية
            self.update_daily_stats(invoice_data['employee_name'])
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"خطأ في إضافة الفاتورة: {e}")
            return False
    
    def update_daily_stats(self, employee_name):
        """تحديث الإحصائيات اليومية"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            today = datetime.now().date()
            
            # حساب إجماليات اليوم
            cursor.execute('''
                SELECT COUNT(*), SUM(quantity), SUM(total_sales)
                FROM invoices 
                WHERE employee_name = ? AND DATE(created_at) = ?
            ''', (employee_name, today))
            
            result = cursor.fetchone()
            total_orders = result[0] or 0
            total_quantity = result[1] or 0
            total_sales = result[2] or 0
            
            # إدراج أو تحديث الإحصائيات
            cursor.execute('''
                INSERT OR REPLACE INTO daily_stats 
                (employee_name, date, total_orders, total_quantity, total_sales)
                VALUES (?, ?, ?, ?, ?)
            ''', (employee_name, today, total_orders, total_quantity, total_sales))
            
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"خطأ في تحديث الإحصائيات: {e}")
    
    def get_employee_monthly_stats(self, employee_name, month=None, year=None):
        """الحصول على إحصائيات الموظف الشهرية"""
        try:
            if month is None:
                month = datetime.now().month
            if year is None:
                year = datetime.now().year
            
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # الحصول على الفواتير للشهر المحدد
            cursor.execute('''
                SELECT * FROM invoices 
                WHERE employee_name = ? 
                AND strftime('%m', created_at) = ? 
                AND strftime('%Y', created_at) = ?
                ORDER BY created_at DESC
            ''', (employee_name, f"{month:02d}", str(year)))
            
            invoices = cursor.fetchall()
            
            # حساب الإجماليات
            total_orders = len(invoices)
            total_quantity = sum(invoice[7] for invoice in invoices)  # quantity column (index 7)
            total_sales = sum(invoice[9] for invoice in invoices)    # total_sales column (index 9)
            
            conn.close()
            
            return {
                'total_orders': total_orders,
                'total_quantity': total_quantity,
                'total_sales': total_sales,
                'invoices': invoices
            }
        except Exception as e:
            print(f"خطأ في الحصول على الإحصائيات: {e}")
            return None
    
    def get_all_invoices_for_shipping(self, days=1, export_type="period", user_id=None):
        """الحصول على جميع الفواتير لملف شركة التوصيل"""
        try:
            conn = self.get_connection()
            
            if export_type == "new_only":
                # الحصول على الطلبات الجديدة فقط (غير مصدرة من قبل)
                query = '''
                    SELECT i.* FROM invoices i
                    LEFT JOIN exported_invoices e ON i.id = e.invoice_id
                    WHERE e.invoice_id IS NULL
                    ORDER BY i.created_at DESC
                '''
            else:
                # الحصول على الفواتير للفترة المحددة
                query = '''
                    SELECT * FROM invoices 
                    WHERE created_at >= datetime('now', '-{} days')
                    ORDER BY created_at DESC
                '''.format(days)
            
            df = pd.read_sql_query(query, conn)
            conn.close()
            
            return df
        except Exception as e:
            print(f"خطأ في الحصول على فواتير التوصيل: {e}")
            return None
    
    def mark_invoices_as_exported(self, invoice_ids, export_type, user_id):
        """تحديد الطلبات كمصدرة"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            for invoice_id in invoice_ids:
                cursor.execute('''
                    INSERT OR IGNORE INTO exported_invoices 
                    (invoice_id, receipt_number, export_type, exported_by)
                    SELECT id, receipt_number, ?, ? FROM invoices WHERE id = ?
                ''', (export_type, user_id, invoice_id))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"خطأ في تحديد الطلبات كمصدرة: {e}")
            return False
    
    def get_export_stats(self):
        """الحصول على إحصائيات التصدير"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # إجمالي الطلبات
            cursor.execute("SELECT COUNT(*) FROM invoices")
            total_invoices = cursor.fetchone()[0]
            
            # الطلبات المصدرة
            cursor.execute("SELECT COUNT(DISTINCT invoice_id) FROM exported_invoices")
            exported_invoices = cursor.fetchone()[0]
            
            # الطلبات الجديدة (غير مصدرة)
            new_invoices = total_invoices - exported_invoices
            
            # آخر تصدير
            cursor.execute("""
                SELECT export_date, export_type, COUNT(*) as count 
                FROM exported_invoices 
                GROUP BY export_date, export_type 
                ORDER BY export_date DESC 
                LIMIT 1
            """)
            last_export = cursor.fetchone()
            
            conn.close()
            
            return {
                'total_invoices': total_invoices,
                'exported_invoices': exported_invoices,
                'new_invoices': new_invoices,
                'last_export': last_export
            }
        except Exception as e:
            print(f"خطأ في الحصول على إحصائيات التصدير: {e}")
            return None
    
    def create_shipping_excel(self, days=1, export_type="period", user_id=None):
        """إنشاء ملف Excel لشركة التوصيل باستخدام القالب الجديد"""
        try:
            df = self.get_all_invoices_for_shipping(days, export_type, user_id)
            if df is None or df.empty:
                print("لا توجد فواتير لإنشاء ملف التوصيل.")
                return None

            # إنشاء ملف إكسل بالشكل الجديد المطلوب
            return self._create_new_shipping_template_excel(df, user_id, export_type)

        except Exception as e:
            print(f"خطأ في إنشاء ملف التوصيل: {e}")
            return None

    def _create_new_shipping_template_excel(self, df, user_id=None, export_type="period"):
        """إنشاء ملف Excel بالشكل الجديد المطلوب"""
        try:
            from openpyxl.styles import Border, Side
            
            # قاموس مفاتيح المحافظات
            GOVERNORATE_CODES = {
                'بغداد': 'BGD',
                'الناصرية ذي قار': 'NAS',
                'ديالى': 'DYL',
                'الكوت واسط': 'KOT',
                'كربلاء': 'KRB',
                'دهوك': 'DOH',
                'بابل الحلة': 'BBL',
                'النجف': 'NJF',
                'البصرة': 'BAS',
                'اربيل': 'ARB',
                'كركوك': 'KRK',
                'السليمانيه': 'SMH',
                'صلاح الدين': 'SAH',
                'الانبار رمادي': 'ANB',
                'السماوة المثنى': 'SAM',
                'موصل': 'MOS',
                'الديوانية': 'DWN',
                'العمارة ميسان': 'AMA'
            }
            
            def get_governorate_code(governorate_name):
                """الحصول على شفرة المحافظة من الاسم"""
                if governorate_name in GOVERNORATE_CODES:
                    return GOVERNORATE_CODES[governorate_name]
                
                for name, code in GOVERNORATE_CODES.items():
                    if governorate_name in name or name in governorate_name:
                        return code
                
                return governorate_name
            
            filename = f"طلبات_التوصيل_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # تحويل البيانات إلى الشكل المطلوب
                converted_data = []
                for index, row in df.iterrows():
                    converted_row = {
                        'ملاحظات': row.get('notes', ''),
                        'عدد القطع\nأجباري': row.get('quantity', 0),
                        'يحتوي على ارجاع بضاعة؟': 'لا',
                        'هاتف المستلم\nأجباري 11 خانة': row.get('client_phone', ''),
                        'تفاصيل العنوان\nأجباري': row.get('nearest_point', ''),
                        'شفرة المحافظة\nأجباري': get_governorate_code(row.get('governorate', '')),
                        'أسم المستلم': row.get('client_name', ''),
                        'المبلغ عراقي\nكامل بالالاف .\nفي حال عدم توفره سيعتبر 0': row.get('total_sales', 0)
                    }
                    converted_data.append(converted_row)
                
                # إنشاء DataFrame جديد
                new_df = pd.DataFrame(converted_data)
                
                # كتابة البيانات الرئيسية
                new_df.to_excel(writer, sheet_name='طلبات الشحن', index=False)
                
                # الحصول على ورقة العمل
                workbook = writer.book
                worksheet = writer.sheets['طلبات الشحن']
                
                # تنسيق العناوين
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # تنسيق حدود العناوين
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # تطبيق التنسيق على العناوين
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # تعديل عرض الأعمدة
                column_widths = {
                    'A': 20,  # ملاحظات
                    'B': 15,  # عدد القطع
                    'C': 25,  # ارجاع بضاعة
                    'D': 25,  # هاتف المستلم
                    'E': 30,  # تفاصيل العنوان
                    'F': 20,  # شفرة المحافظة
                    'G': 20,  # أسم المستلم
                    'H': 35   # المبلغ عراقي
                }
                
                for col_letter, width in column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width
                
                # تنسيق البيانات
                data_alignment = Alignment(horizontal="center", vertical="center")
                data_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # تطبيق التنسيق على البيانات
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = data_alignment
                        cell.border = data_border
                
                # إنشاء ورقة مفاتيح المحافظات
                governorate_sheet = workbook.create_sheet("مفاتيح المحافظات")
                
                # إضافة مفاتيح المحافظات
                governorate_data = []
                for governorate, code in GOVERNORATE_CODES.items():
                    governorate_data.append([governorate, code])
                
                governorate_df = pd.DataFrame(governorate_data, columns=['المحافظة', 'الشفرة'])
                governorate_df.to_excel(writer, sheet_name='مفاتيح المحافظات', index=False)
                
                # تنسيق ورقة مفاتيح المحافظات
                gov_worksheet = writer.sheets['مفاتيح المحافظات']
                
                # تنسيق العناوين
                for cell in gov_worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # تعديل عرض الأعمدة
                gov_worksheet.column_dimensions['A'].width = 25
                gov_worksheet.column_dimensions['B'].width = 15
                
                # تنسيق البيانات
                for row in gov_worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = data_alignment
                        cell.border = data_border
            
            # تحديد الطلبات كمصدرة إذا كان المستخدم محدد
            if user_id and not df.empty:
                invoice_ids = df['id'].tolist()
                self.mark_invoices_as_exported(invoice_ids, export_type, user_id)
            
            print(f"تم إنشاء ملف التوصيل بالشكل الجديد بنجاح: {filename}")
            return filename
            
        except Exception as e:
            print(f"خطأ في إنشاء ملف التوصيل بالشكل الجديد: {e}")
            return None

    def _create_new_shipping_excel(self, df, user_id=None, export_type="period"):
        """إنشاء ملف Excel جديد في حالة عدم وجود القالب (للتوافق)"""
        return self._create_new_shipping_template_excel(df, user_id, export_type)
    
    def add_user(self, telegram_id, username, full_name, role="employee"):
        """إضافة مستخدم جديد"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT OR REPLACE INTO users (telegram_id, username, full_name, role)
                VALUES (?, ?, ?, ?)
            ''', (telegram_id, username, full_name, role))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"خطأ في إضافة المستخدم: {e}")
            return False
    
    def get_user_role(self, telegram_id):
        """الحصول على دور المستخدم"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT role FROM users WHERE telegram_id = ?', (telegram_id,))
            result = cursor.fetchone()
            
            conn.close()
            return result[0] if result else "employee"
        except Exception as e:
            print(f"خطأ في الحصول على دور المستخدم: {e}")
            return "employee"
    
    def get_all_employees(self):
        """الحصول على قائمة جميع الموظفين"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # الحصول على الموظفين من جدول users أولاً
            cursor.execute('''
                SELECT full_name FROM users 
                WHERE role IN ('employee', 'warehouse_manager')
                ORDER BY full_name
            ''')
            
            users_results = cursor.fetchall()
            
            # الحصول على الموظفين من جدول invoices أيضاً (للتوافق مع النظام القديم)
            cursor.execute('''
                SELECT DISTINCT employee_name FROM invoices 
                ORDER BY employee_name
            ''')
            
            invoices_results = cursor.fetchall()
            
            conn.close()
            
            # دمج النتائج وإزالة التكرار
            all_employees = set()
            
            # إضافة الموظفين من جدول users
            for row in users_results:
                if row[0]:  # التأكد من أن الاسم ليس فارغاً
                    all_employees.add(row[0])
            
            # إضافة الموظفين من جدول invoices
            for row in invoices_results:
                if row[0]:  # التأكد من أن الاسم ليس فارغاً
                    all_employees.add(row[0])
            
            return sorted(list(all_employees))
        except Exception as e:
            print(f"خطأ في الحصول على قائمة الموظفين: {e}")
            return []
    
    def get_employees_with_passwords(self):
        """الحصول على قائمة الموظفين الذين لديهم كلمات مرور"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT employee_name FROM employee_passwords 
                ORDER BY employee_name
            ''')
            
            results = cursor.fetchall()
            conn.close()
            
            return [row[0] for row in results]
        except Exception as e:
            print(f"خطأ في الحصول على قائمة الموظفين الذين لديهم كلمات مرور: {e}")
            return []
    
    def get_all_employees_stats(self, month=None, year=None):
        """الحصول على إحصائيات جميع الموظفين"""
        try:
            if month is None:
                month = datetime.now().month
            if year is None:
                year = datetime.now().year
            
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # الحصول على جميع الموظفين أولاً
            all_employees = self.get_all_employees()
            
            # الحصول على إحصائيات الموظفين الذين لديهم فواتير
            cursor.execute('''
                SELECT 
                    employee_name,
                    COUNT(*) as total_orders,
                    SUM(quantity) as total_quantity,
                    SUM(total_sales) as total_sales
                FROM invoices 
                WHERE strftime('%m', created_at) = ? 
                AND strftime('%Y', created_at) = ?
                GROUP BY employee_name
            ''', (f"{month:02d}", str(year)))
            
            stats_results = cursor.fetchall()
            conn.close()
            
            # إنشاء قاموس للإحصائيات
            stats_dict = {}
            for row in stats_results:
                stats_dict[row[0]] = {
                    'total_orders': row[1],
                    'total_quantity': row[2] or 0,
                    'total_sales': row[3] or 0
                }
            
            # إنشاء قائمة شاملة لجميع الموظفين مع إحصائياتهم
            all_stats = []
            for employee in all_employees:
                if employee in stats_dict:
                    # الموظف لديه إحصائيات
                    stats = stats_dict[employee]
                    all_stats.append((
                        employee,
                        stats['total_orders'],
                        stats['total_quantity'],
                        stats['total_sales']
                    ))
                else:
                    # الموظف ليس لديه إحصائيات (صفر)
                    all_stats.append((employee, 0, 0, 0))
            
            # ترتيب حسب إجمالي المبيعات (تنازلي)
            all_stats.sort(key=lambda x: x[3], reverse=True)
            
            return all_stats
        except Exception as e:
            print(f"خطأ في الحصول على إحصائيات الموظفين: {e}")
            return []

    def get_employee_stats_by_date_range(self, employee_name, start_date, end_date):
        """الحصول على إحصائيات الموظف لفترة زمنية محددة"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT * FROM invoices 
                WHERE employee_name = ? 
                AND DATE(created_at) BETWEEN ? AND ?
                ORDER BY created_at DESC
            ''', (employee_name, start_date, end_date))
            
            invoices = cursor.fetchall()
            
            # حساب الإجماليات
            total_orders = len(invoices)
            total_quantity = sum(invoice[7] for invoice in invoices)  # quantity column
            total_sales = sum(invoice[9] for invoice in invoices)    # total_sales column
            
            conn.close()
            
            return {
                'total_orders': total_orders,
                'total_quantity': total_quantity,
                'total_sales': total_sales,
                'invoices': invoices,
                'start_date': start_date,
                'end_date': end_date
            }
        except Exception as e:
            print(f"خطأ في الحصول على الإحصائيات: {e}")
            return None

    def get_all_employees_stats_by_date_range(self, start_date, end_date):
        """الحصول على إحصائيات جميع الموظفين لفترة زمنية محددة"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # الحصول على جميع الموظفين أولاً
            all_employees = self.get_all_employees()
            
            # الحصول على إحصائيات الموظفين الذين لديهم فواتير
            cursor.execute('''
                SELECT 
                    employee_name,
                    COUNT(*) as total_orders,
                    SUM(quantity) as total_quantity,
                    SUM(total_sales) as total_sales
                FROM invoices 
                WHERE DATE(created_at) BETWEEN ? AND ?
                GROUP BY employee_name
            ''', (start_date, end_date))
            
            stats_results = cursor.fetchall()
            conn.close()
            
            # إنشاء قاموس للإحصائيات
            stats_dict = {}
            for row in stats_results:
                stats_dict[row[0]] = {
                    'total_orders': row[1],
                    'total_quantity': row[2] or 0,
                    'total_sales': row[3] or 0
                }
            
            # إنشاء قائمة شاملة لجميع الموظفين مع إحصائياتهم
            all_stats = []
            for employee in all_employees:
                if employee in stats_dict:
                    # الموظف لديه إحصائيات
                    stats = stats_dict[employee]
                    all_stats.append((
                        employee,
                        stats['total_orders'],
                        stats['total_quantity'],
                        stats['total_sales']
                    ))
                else:
                    # الموظف ليس لديه إحصائيات (صفر)
                    all_stats.append((employee, 0, 0, 0))
            
            # ترتيب حسب إجمالي المبيعات (تنازلي)
            all_stats.sort(key=lambda x: x[3], reverse=True)
            
            return all_stats
        except Exception as e:
            print(f"خطأ في الحصول على إحصائيات الموظفين: {e}")
            return []

    def create_statistics_excel(self, stats_data, report_type="monthly", date_info=""):
        """إنشاء ملف إكسل للإحصائيات"""
        try:
            # إنشاء DataFrame
            if report_type == "all_employees":
                df = pd.DataFrame(stats_data['employees'])
                filename = f"employees_statistics_{date_info}_{int(time.time())}.xlsx"
            else:
                df = pd.DataFrame([stats_data])
                filename = f"employee_statistics_{stats_data.get('employee_name', 'unknown')}_{date_info}_{int(time.time())}.xlsx"
            
            # إنشاء ملف الإكسل
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Statistics', index=False)
                
                # الحصول على ورقة العمل
                worksheet = writer.sheets['Statistics']
                
                # تنسيق العناوين
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # تنسيق الأرقام
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal="center")
            
            return filename
            
        except Exception as e:
            print(f"خطأ في إنشاء ملف الإكسل: {e}")
            return None

    def set_employee_password(self, employee_name, password):
        """تعيين كلمة مرور للموظف"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT OR REPLACE INTO employee_passwords (employee_name, password, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
            ''', (employee_name, password))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"خطأ في تعيين كلمة المرور: {e}")
            return False

    def verify_employee_password(self, employee_name, password):
        """التحقق من كلمة مرور الموظف"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT password FROM employee_passwords 
                WHERE employee_name = ?
            ''', (employee_name,))
            
            result = cursor.fetchone()
            conn.close()
            
            if result:
                return result[0] == password
            return False
        except Exception as e:
            print(f"خطأ في التحقق من كلمة المرور: {e}")
            return False

    def get_employee_password(self, employee_name):
        """الحصول على كلمة مرور الموظف"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT password FROM employee_passwords 
                WHERE employee_name = ?
            ''', (employee_name,))
            
            result = cursor.fetchone()
            conn.close()
            
            return result[0] if result else None
        except Exception as e:
            print(f"خطأ في الحصول على كلمة المرور: {e}")
            return None

    def has_password(self, employee_name):
        """التحقق من وجود كلمة مرور للموظف"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT COUNT(*) FROM employee_passwords 
                WHERE employee_name = ?
            ''', (employee_name,))
            
            result = cursor.fetchone()
            conn.close()
            
            return result[0] > 0
        except Exception as e:
            print(f"خطأ في التحقق من وجود كلمة المرور: {e}")
            return False

    def get_all_passwords(self):
        """الحصول على جميع كلمات المرور"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT employee_name, password FROM employee_passwords 
                ORDER BY employee_name
            ''')
            
            result = cursor.fetchall()
            conn.close()
            
            return result
        except Exception as e:
            print(f"خطأ في الحصول على كلمات المرور: {e}")
            return []

    def delete_employee_password(self, employee_name):
        """حذف كلمة مرور الموظف"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                DELETE FROM employee_passwords 
                WHERE employee_name = ?
            ''', (employee_name,))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"خطأ في حذف كلمة المرور: {e}")
            return False

    # ==================== دوال إدارة المرتجعات ====================
    
    def get_invoice_by_receipt(self, receipt_number):
        """الحصول على فاتورة برقم الإيصال"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # تنظيف رقم الإيصال من المسافات
            receipt_number = receipt_number.strip()
            
            # إذا كان الرقم بدون بادئة INV-، أضفها
            if not receipt_number.startswith('INV-'):
                receipt_number = f"INV-{receipt_number}"
            
            cursor.execute('''
                SELECT id, receipt_number, employee_name, client_name, 
                       quantity, price, total_sales, created_at
                FROM invoices 
                WHERE receipt_number = ?
            ''', (receipt_number,))
            
            result = cursor.fetchone()
            conn.close()
            
            if result:
                return {
                    'id': result[0],
                    'receipt_number': result[1],
                    'employee_name': result[2],
                    'client_name': result[3],
                    'quantity': result[4],
                    'price': result[5],
                    'total_sales': result[6],
                    'created_at': result[7]
                }
            return None
        except Exception as e:
            print(f"خطأ في الحصول على الفاتورة: {e}")
            return None

    def add_return(self, return_data):
        """إضافة مرتجع جديد"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO returns (
                    invoice_id, receipt_number, employee_name, return_type,
                    returned_quantity, returned_amount, remaining_amount,
                    return_reason, processed_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                return_data['invoice_id'],
                return_data['receipt_number'],
                return_data['employee_name'],
                return_data['return_type'],
                return_data['returned_quantity'],
                return_data['returned_amount'],
                return_data['remaining_amount'],
                return_data['return_reason'],
                return_data['processed_by']
            ))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"خطأ في إضافة المرتجع: {e}")
            return False

    def get_returns_by_employee(self, employee_name, start_date=None, end_date=None):
        """الحصول على مرتجعات موظف معين"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            query = '''
                SELECT r.*, i.client_name, i.quantity as original_quantity, i.total_sales as original_total
                FROM returns r
                JOIN invoices i ON r.invoice_id = i.id
                WHERE r.employee_name = ?
            '''
            params = [employee_name]
            
            if start_date and end_date:
                query += ' AND DATE(r.created_at) BETWEEN ? AND ?'
                params.extend([start_date, end_date])
            
            query += ' ORDER BY r.created_at DESC'
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            returns = []
            for row in results:
                returns.append({
                    'id': row[0],
                    'invoice_id': row[1],
                    'receipt_number': row[2],
                    'employee_name': row[3],
                    'return_type': row[4],
                    'returned_quantity': row[5],
                    'returned_amount': row[6],
                    'remaining_amount': row[7],
                    'return_reason': row[8],
                    'processed_by': row[9],
                    'created_at': row[10],
                    'client_name': row[11],
                    'original_quantity': row[12],
                    'original_total': row[13]
                })
            
            return returns
        except Exception as e:
            print(f"خطأ في الحصول على المرتجعات: {e}")
            return []

    def get_employee_stats_with_returns(self, employee_name, month=None, year=None):
        """الحصول على إحصائيات الموظف مع خصم المرتجعات"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # الحصول على إحصائيات الفواتير
            if month and year:
                cursor.execute('''
                    SELECT COUNT(*) as total_invoices,
                           SUM(quantity) as total_quantity,
                           SUM(total_sales) as total_sales
                    FROM invoices 
                    WHERE employee_name = ? 
                    AND strftime('%m', created_at) = ? 
                    AND strftime('%Y', created_at) = ?
                ''', (employee_name, f"{month:02d}", str(year)))
            else:
                current_month = datetime.now().month
                current_year = datetime.now().year
                cursor.execute('''
                    SELECT COUNT(*) as total_invoices,
                           SUM(quantity) as total_quantity,
                           SUM(total_sales) as total_sales
                    FROM invoices 
                    WHERE employee_name = ? 
                    AND strftime('%m', created_at) = ? 
                    AND strftime('%Y', created_at) = ?
                ''', (employee_name, f"{current_month:02d}", str(current_year)))
            
            invoice_stats = cursor.fetchone()
            
            # الحصول على إحصائيات المرتجعات
            if month and year:
                cursor.execute('''
                    SELECT SUM(returned_quantity) as total_returned_quantity,
                           SUM(returned_amount) as total_returned_amount
                    FROM returns 
                    WHERE employee_name = ? 
                    AND strftime('%m', created_at) = ? 
                    AND strftime('%Y', created_at) = ?
                ''', (employee_name, f"{month:02d}", str(year)))
            else:
                current_month = datetime.now().month
                current_year = datetime.now().year
                cursor.execute('''
                    SELECT SUM(returned_quantity) as total_returned_quantity,
                           SUM(returned_amount) as total_returned_amount
                    FROM returns 
                    WHERE employee_name = ? 
                    AND strftime('%m', created_at) = ? 
                    AND strftime('%Y', created_at) = ?
                ''', (employee_name, f"{current_month:02d}", str(current_year)))
            
            return_stats = cursor.fetchone()
            
            conn.close()
            
            # حساب الإحصائيات النهائية
            total_invoices = invoice_stats[0] or 0
            total_quantity = invoice_stats[1] or 0
            total_sales = invoice_stats[2] or 0
            
            total_returned_quantity = return_stats[0] or 0
            total_returned_amount = return_stats[1] or 0
            
            # خصم المرتجعات
            final_quantity = total_quantity - total_returned_quantity
            final_sales = total_sales - total_returned_amount
            
            return {
                'employee_name': employee_name,
                'total_invoices': total_invoices,
                'total_quantity': total_quantity,
                'total_sales': total_sales,
                'returned_quantity': total_returned_quantity,
                'returned_amount': total_returned_amount,
                'final_quantity': final_quantity,
                'final_sales': final_sales
            }
        except Exception as e:
            print(f"خطأ في الحصول على الإحصائيات مع المرتجعات: {e}")
            return None

    def get_all_employees_stats_with_returns(self, month=None, year=None):
        """الحصول على إحصائيات جميع الموظفين مع خصم المرتجعات"""
        try:
            # الحصول على جميع الموظفين من دالة get_all_employees
            all_employees = self.get_all_employees()
            
            # الحصول على إحصائيات كل موظف
            all_stats = []
            for employee in all_employees:
                stats = self.get_employee_stats_with_returns(employee, month, year)
                if stats:
                    all_stats.append(stats)
                else:
                    # إذا لم تكن هناك إحصائيات، إضافة إحصائيات فارغة
                    all_stats.append({
                        'employee_name': employee,
                        'total_invoices': 0,
                        'total_quantity': 0,
                        'total_sales': 0,
                        'returned_quantity': 0,
                        'returned_amount': 0,
                        'final_quantity': 0,
                        'final_sales': 0
                    })
            
            # ترتيب حسب المبيعات النهائية (تنازلي)
            all_stats.sort(key=lambda x: x['final_sales'], reverse=True)
            
            return all_stats
        except Exception as e:
            print(f"خطأ في الحصول على إحصائيات جميع الموظفين: {e}")
            return [] 